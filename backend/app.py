from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
import sys
import pandas as pd
import tempfile
import json
from openpyxl import load_workbook
from datetime import datetime
import numpy as np
import uuid

app = Flask(__name__)
CORS(app)  # Enable CORS for all origins (production-ready)

# Environment configuration
app.config['DEBUG'] = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
app.config['HOST'] = os.environ.get('FLASK_HOST', '0.0.0.0')
app.config['PORT'] = int(os.environ.get('PORT', 5001))  # Dynamic port for production

# Configuration
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB

# Create directory for storing original uploaded files
UPLOAD_STORAGE_DIR = 'uploaded_files'
if not os.path.exists(UPLOAD_STORAGE_DIR):
    os.makedirs(UPLOAD_STORAGE_DIR)

# Create directory for temporary JSON files
JSON_TEMP_DIR = 'temp_json_files'
if not os.path.exists(JSON_TEMP_DIR):
    os.makedirs(JSON_TEMP_DIR)

# Removed convert_to_json_serializable - no longer converting files to JSON

def save_uploaded_file(file_data, filename, is_schema=False, box_number=1):
    """Save uploaded file with original filename in organized subdirectories"""
    # Generate unique ID for this file
    unique_id = str(uuid.uuid4())
    
    # Create subdirectory based on box number
    subdirectory = f"bank{box_number}"
    bank_dir = os.path.join(UPLOAD_STORAGE_DIR, subdirectory)
    
    # Create the subdirectory if it doesn't exist
    if not os.path.exists(bank_dir):
        os.makedirs(bank_dir)
    
    # Use exact original filename for main.py compatibility
    saved_filename = filename
    saved_filepath = os.path.join(bank_dir, saved_filename)
    
    # If file already exists, remove it first to prevent duplicates
    if os.path.exists(saved_filepath):
        try:
            os.remove(saved_filepath)
            print(f"Removed existing file: {saved_filename}")
        except Exception as e:
            print(f"Warning: Could not remove existing file {saved_filename}: {e}")
            # If we can't remove it, use timestamp-based naming as fallback
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = os.path.splitext(filename)[0]
            extension = os.path.splitext(filename)[1]
            saved_filename = f"{base_name}_{timestamp}{extension}"
            saved_filepath = os.path.join(bank_dir, saved_filename)
    
    # Save the original file
    file_data.save(saved_filepath)
    
    return saved_filename, unique_id, subdirectory

def cleanup_temp_json_files():
    """Clean up temporary JSON files"""
    try:
        if os.path.exists(JSON_TEMP_DIR):
            for filename in os.listdir(JSON_TEMP_DIR):
                if filename.endswith('.json'):
                    file_path = os.path.join(JSON_TEMP_DIR, filename)
                    os.remove(file_path)
                    print(f"Cleaned up JSON: {filename}")
        return True
    except Exception as e:
        print(f"Error cleaning up JSON files: {str(e)}")
        return False

def cleanup_uploaded_files():
    """Clean up all uploaded XLSX/CSV files from bank1 and bank2"""
    try:
        cleaned_files = []
        failed_files = []
        
        # Force garbage collection to close any open file handles
        import gc
        gc.collect()
        
        # Clean bank1 directory
        bank1_dir = os.path.join(UPLOAD_STORAGE_DIR, 'bank1')
        if os.path.exists(bank1_dir):
            for filename in os.listdir(bank1_dir):
                if filename.endswith(('.xlsx', '.csv', '.xls')):
                    file_path = os.path.join(bank1_dir, filename)
                    try:
                        # Try to remove the file
                        os.remove(file_path)
                        cleaned_files.append(f"bank1/{filename}")
                        print(f"Cleaned up uploaded file: bank1/{filename}")
                    except PermissionError:
                        # Try to force remove with different approach
                        try:
                            import shutil
                            shutil.rmtree(file_path, ignore_errors=True)
                            cleaned_files.append(f"bank1/{filename} (forced)")
                            print(f"Force cleaned uploaded file: bank1/{filename}")
                        except:
                            failed_files.append(f"bank1/{filename} (file in use)")
                            print(f"Failed to clean: bank1/{filename} (file in use)")
                    except Exception as e:
                        failed_files.append(f"bank1/{filename} ({str(e)})")
                        print(f"Failed to clean: bank1/{filename} - {str(e)}")
        
        # Clean bank2 directory
        bank2_dir = os.path.join(UPLOAD_STORAGE_DIR, 'bank2')
        if os.path.exists(bank2_dir):
            for filename in os.listdir(bank2_dir):
                if filename.endswith(('.xlsx', '.csv', '.xls')):
                    file_path = os.path.join(bank2_dir, filename)
                    try:
                        # Try to remove the file
                        os.remove(file_path)
                        cleaned_files.append(f"bank2/{filename}")
                        print(f"Cleaned up uploaded file: bank2/{filename}")
                    except PermissionError:
                        # Try to force remove with different approach
                        try:
                            import shutil
                            shutil.rmtree(file_path, ignore_errors=True)
                            cleaned_files.append(f"bank2/{filename} (forced)")
                            print(f"Force cleaned uploaded file: bank2/{filename}")
                        except:
                            failed_files.append(f"bank2/{filename} (file in use)")
                            print(f"Failed to clean: bank2/{filename} (file in use)")
                    except Exception as e:
                        failed_files.append(f"bank2/{filename} ({str(e)})")
                        print(f"Failed to clean: bank2/{filename} - {str(e)}")
        
        return len(failed_files) == 0, cleaned_files, failed_files
    except Exception as e:
        print(f"Error cleaning up uploaded files: {str(e)}")
        return False, [], [f"General error: {str(e)}"]

def cleanup_all_files():
    """Clean up both JSON files and uploaded files"""
    try:
        # Clean JSON files
        json_success = cleanup_temp_json_files()
        
        # Clean uploaded files
        upload_success, cleaned_files, failed_files = cleanup_uploaded_files()
        
        # Combine all cleaned files
        all_cleaned_files = cleaned_files.copy()
        
        # Add JSON files to the list if they exist
        if os.path.exists(JSON_TEMP_DIR):
            json_files = [f for f in os.listdir(JSON_TEMP_DIR) if f.endswith('.json')]
            all_cleaned_files.extend([f"temp_json/{f}" for f in json_files])
        
        return upload_success, all_cleaned_files, failed_files
    except Exception as e:
        print(f"Error in cleanup_all_files: {str(e)}")
        return False, [], [f"General error: {str(e)}"]

def create_fallback_excel_file(json_files, output_path):
    """Create a fallback Excel file with all JSON data combined"""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Combined Data"
        
        # Track all data
        all_data = []
        sheet_names = []
        
        # Process each JSON file
        for json_file in json_files:
            if os.path.exists(json_file):
                try:
                    with open(json_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    
                    # Extract sheet names and data
                    for sheet_name, sheet_data in data.items():
                        if sheet_name.startswith('_') or sheet_name.lower() in ['metadata', 'data']:
                            continue
                        
                        if isinstance(sheet_data, list) and len(sheet_data) > 0:
                            # Convert to DataFrame
                            df = pd.DataFrame(sheet_data)
                            
                            # Add source information
                            df['source_file'] = os.path.basename(json_file)
                            df['source_sheet'] = sheet_name
                            
                            all_data.append(df)
                            sheet_names.append(f"{os.path.basename(json_file)}_{sheet_name}")
                            
                except Exception as e:
                    print(f"Error processing {json_file}: {e}")
        
        if all_data:
            # Combine all data
            combined_df = pd.concat(all_data, ignore_index=True, sort=False)
            
            # Write to Excel
            combined_df.to_excel(output_path, index=False, sheet_name="Combined Data")
            
            print(f"Created fallback Excel file with {len(combined_df)} rows from {len(all_data)} data sources")
            return output_path
        else:
            print("No data found to create Excel file")
            return None
            
    except Exception as e:
        print(f"Error creating fallback Excel file: {e}")
        return None

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_schema_title_from_excel(file_path, sheet_name):
    """Extract title from rows 1-4, columns A-B in Excel sheet for schema files"""
    try:
        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook[sheet_name]
        
        # Check merged cells first
        for merged_range in worksheet.merged_cells.ranges:
            if merged_range.min_row <= 4 and merged_range.max_row >= 1 and \
               merged_range.min_col <= 2 and merged_range.max_col >= 1:
                title_cell = worksheet.cell(merged_range.min_row, merged_range.min_col)
                if title_cell.value:
                    return str(title_cell.value).strip()
        
        # Try to get title from cells A1:B4
        title_parts = []
        for row in range(1, 5):  # Rows 1-4
            for col in range(1, 3):  # Columns A-B
                cell_value = worksheet.cell(row, col).value
                if cell_value:
                    title_parts.append(str(cell_value).strip())
        
        if title_parts:
            return ' '.join(title_parts)
        
        return f"Tab: {sheet_name}"  # Default title
        
    except Exception as e:
        print(f"Error extracting schema title from {sheet_name}: {str(e)}")
        return f"Tab: {sheet_name}"

# Removed convert_csv_to_json - no longer converting files to JSON

# Removed convert_excel_to_json - no longer converting files to JSON

# Removed convert_file_to_json - no longer converting files to JSON

@app.route('/api/download-files', methods=['POST'])
def download_files():
    print("Downloading files")
    

@app.route('/api/process-files', methods=['POST'])
def process_files():
    """Process uploaded files and save them for main.py access"""
    try:
        if 'files' not in request.files:
            return jsonify({'error': 'No files uploaded'}), 400
        
        # Check if this is a schema file request
        is_schema = request.args.get('schema', 'false').lower() == 'true'
        
        # Get box number (1 or 2) to determine subdirectory
        box_number = int(request.args.get('box', 1))
        
        files = request.files.getlist('files')
        
        # Filter out empty files
        valid_files = [file for file in files if file and file.filename != '']
        
        if len(valid_files) == 0:
            return jsonify({'error': 'Please upload at least 1 file'}), 400
        
        results = []
        
        for file in valid_files:
            if allowed_file(file.filename):
                # Check file size
                file.seek(0, 2)  # Seek to end
                file_size = int(file.tell())
                file.seek(0)  # Reset to beginning
                
                if file_size > MAX_FILE_SIZE:
                    results.append({
                        'filename': file.filename,
                        'error': True,
                        'message': f"File too large (max {MAX_FILE_SIZE // (1024*1024)}MB)"
                    })
                    continue
                
                # Reset file pointer to beginning for saving
                file.seek(0)
                
                # Save the original file
                saved_filename = None
                unique_id = None
                try:
                    saved_filename, unique_id, subdirectory = save_uploaded_file(file, file.filename, is_schema, box_number)
                    print(f"DEBUG: Saved original file {file.filename} as {saved_filename}")
                    
                    results.append({
                        'filename': file.filename,
                        'saved_filename': saved_filename,
                        'unique_id': unique_id,
                        'file_size': file_size,
                        'file_type': 'schema' if is_schema else 'data',
                        'subdirectory': subdirectory,
                        'error': False
                    })
                    
                except Exception as save_error:
                    results.append({
                        'filename': file.filename,
                        'error': True,
                        'message': f"Error saving file: {str(save_error)}"
                    })
            
            else:
                results.append({
                    'filename': file.filename,
                    'error': True,
                    'message': "Invalid file type. Only CSV and Excel files are supported."
                })
        
        print(f"DEBUG: Final results count: {len(results)}")  # Debug log
        
        return jsonify({
            'success': True,
            'results': results,
            'file_count': len(valid_files),
            'is_schema': is_schema
        })
    
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({'status': 'healthy', 'message': 'Bridgette backend is running'})

@app.route('/api/uploaded-files', methods=['GET'])
def list_uploaded_files():
    """List all saved uploaded files"""
    try:
        if not os.path.exists(UPLOAD_STORAGE_DIR):
            return jsonify({'files': []})
        
        files = []
        for filename in os.listdir(UPLOAD_STORAGE_DIR):
            if filename.endswith(('.csv', '.xlsx', '.xls')):
                filepath = os.path.join(UPLOAD_STORAGE_DIR, filename)
                file_stats = os.stat(filepath)
                files.append({
                    'filename': filename,
                    'size': file_stats.st_size,
                    'created': datetime.fromtimestamp(file_stats.st_ctime).isoformat(),
                    'modified': datetime.fromtimestamp(file_stats.st_mtime).isoformat()
                })
        
        # Sort by creation time (newest first)
        files.sort(key=lambda x: x['created'], reverse=True)
        
        return jsonify({'files': files})
    
    except Exception as e:
        return jsonify({'error': f'Error listing files: {str(e)}'}), 500

@app.route('/api/uploaded-files/<filename>', methods=['GET'])
def get_uploaded_file(filename):
    """Get a specific uploaded file by filename"""
    try:
        filepath = os.path.join(UPLOAD_STORAGE_DIR, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        # Return file info instead of content
        file_stats = os.stat(filepath)
        return jsonify({
            'filename': filename,
            'size': file_stats.st_size,
            'created': datetime.fromtimestamp(file_stats.st_ctime).isoformat(),
            'modified': datetime.fromtimestamp(file_stats.st_mtime).isoformat(),
            'filepath': filepath
        })
    
    except Exception as e:
        return jsonify({'error': f'Error reading file: {str(e)}'}), 500

@app.route('/api/uploaded-files/<filename>', methods=['DELETE'])
def delete_uploaded_file(filename):
    """Delete a specific uploaded file"""
    try:
        filepath = os.path.join(UPLOAD_STORAGE_DIR, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        os.remove(filepath)
        return jsonify({'message': f'File {filename} deleted successfully'})
    
    except Exception as e:
        return jsonify({'error': f'Error deleting file: {str(e)}'}), 500

# Removed start-merging endpoint - files are processed directly by main.py

@app.route('/api/process-with-main', methods=['POST'])
def process_with_main():
    """Process JSON files using main.py functionality"""
    try:
        data = request.get_json()
        
        if not data or 'action' not in data:
            return jsonify({'error': 'No action specified'}), 400
        
        action = data['action']
        
        if action == 'analyze_schemas':
            return analyze_schemas_with_main()
        elif action == 'compare_schemas':
            return compare_schemas_with_main()
        elif action == 'process_file':
            file_path = data.get('file_path')
            if not file_path:
                return jsonify({'error': 'No file path provided'}), 400
            return process_file_with_main(file_path)
        else:
            return jsonify({'error': f'Unknown action: {action}'}), 400
    
    except Exception as e:
        return jsonify({'error': f'Error processing with main.py: {str(e)}'}), 500

def analyze_schemas_with_main():
    """Process uploaded XLSX/CSV files using main.py and analyze schemas"""
    try:
        # Import main.py functions
        sys.path.append(os.path.dirname(os.path.abspath(__file__)))
        from main import process_file, count_schemas_in_json
        
        # Find XLSX/CSV files in bank1 and bank2 directories
        bank1_files = []
        bank2_files = []
        
        bank1_dir = os.path.join(UPLOAD_STORAGE_DIR, 'bank1')
        bank2_dir = os.path.join(UPLOAD_STORAGE_DIR, 'bank2')
        
        if os.path.exists(bank1_dir):
            bank1_files = [f for f in os.listdir(bank1_dir) if f.lower().endswith(('.xlsx', '.csv', '.xls'))]
        if os.path.exists(bank2_dir):
            bank2_files = [f for f in os.listdir(bank2_dir) if f.lower().endswith(('.xlsx', '.csv', '.xls'))]
        
        if len(bank1_files) == 0 and len(bank2_files) == 0:
            return jsonify({'error': 'No XLSX/CSV files found in uploaded_files directories'}), 400
        
        processed_files = []
        json_files = []
        
        # Process bank1 files
        for filename in bank1_files:
            file_path = os.path.join(bank1_dir, filename)
            try:
                # Process file with main.py - this creates JSON files
                result = process_file(file_path, clean_data=True, include_metadata=True)
                processed_files.append({
                    'original_file': filename,
                    'directory': 'bank1',
                    'json_file': result.get('output_file', f"{os.path.splitext(filename)[0]}_converted.json")
                })
                json_files.append(result.get('output_file', f"{os.path.splitext(filename)[0]}_converted.json"))
            except Exception as e:
                print(f"Error processing {filename}: {str(e)}")
        
        # Process bank2 files
        for filename in bank2_files:
            file_path = os.path.join(bank2_dir, filename)
            try:
                # Process file with main.py - this creates JSON files
                result = process_file(file_path, clean_data=True, include_metadata=True)
                processed_files.append({
                    'original_file': filename,
                    'directory': 'bank2',
                    'json_file': result.get('output_file', f"{os.path.splitext(filename)[0]}_converted.json")
                })
                json_files.append(result.get('output_file', f"{os.path.splitext(filename)[0]}_converted.json"))
            except Exception as e:
                print(f"Error processing {filename}: {str(e)}")
        
        # Count schemas in JSON files
        schema_counts = []
        for json_file in json_files:
            if os.path.exists(json_file):
                try:
                    count = count_schemas_in_json(json_file)
                    schema_counts.append({'file': json_file, 'count': count})
                except Exception as e:
                    print(f"Error counting schemas in {json_file}: {str(e)}")
        
        return jsonify({
            'success': True,
            'message': 'Files processed successfully with main.py',
            'processed_files': processed_files,
            'json_files_created': json_files,
            'schema_counts': schema_counts,
            'total_schemas': sum(sc['count'] for sc in schema_counts)
        })
    
    except Exception as e:
        return jsonify({'error': f'Error processing files with main.py: {str(e)}'}), 500

@app.route('/api/trigger-main-processing', methods=['POST'])
def trigger_main_processing():
    """Process all uploaded XLSX/CSV files with main.py and return results"""
    try:
        # Temporarily redirect stdout to suppress Unicode print statements
        from io import StringIO
        import time
        import json
        old_stdout = sys.stdout
        old_stderr = sys.stderr
        sys.stdout = StringIO()
        sys.stderr = StringIO()
        
        # Import main.py functions
        sys.path.append(os.path.dirname(os.path.abspath(__file__)))
        from main import process_file, count_schemas_in_json, send_json_to_chatgpt, process_chatgpt_schema_analysis, create_combined_customer_data
        
        # Find all XLSX/CSV files in uploaded_files directories
        all_files = []
        
        # Check bank1 directory
        bank1_dir = os.path.join(UPLOAD_STORAGE_DIR, 'bank1')
        if os.path.exists(bank1_dir):
            for filename in os.listdir(bank1_dir):
                if filename.lower().endswith(('.xlsx', '.csv', '.xls')):
                    all_files.append({
                        'file': filename,
                        'path': os.path.join(bank1_dir, filename),
                        'directory': 'bank1'
                    })
        
        # Check bank2 directory
        bank2_dir = os.path.join(UPLOAD_STORAGE_DIR, 'bank2')
        if os.path.exists(bank2_dir):
            for filename in os.listdir(bank2_dir):
                if filename.lower().endswith(('.xlsx', '.csv', '.xls')):
                    all_files.append({
                        'file': filename,
                        'path': os.path.join(bank2_dir, filename),
                        'directory': 'bank2'
                    })
        
        if len(all_files) == 0:
            return jsonify({
                'success': False,
                'message': 'No XLSX/CSV files found in uploaded_files directories',
                'files_found': 0
            })
        
        # Process each file with main.py
        results = []
        json_files_created = []
        
        for file_info in all_files:
            try:
                print(f"Processing {file_info['file']} from {file_info['directory']}")
                
                # Create output file in temp directory
                base_name = os.path.splitext(file_info['file'])[0]
                temp_output_file = os.path.join(JSON_TEMP_DIR, f"{base_name}_converted.json")
                
                # Process file with better error handling
                try:
                    result = process_file(file_info['path'], output_file=temp_output_file, clean_data=True, include_metadata=True)
                    
                    # Check if result is valid
                    if result is None:
                        raise Exception("process_file returned None")
                    
                    results.append({
                        'original_file': file_info['file'],
                        'directory': file_info['directory'],
                        'json_file': temp_output_file,
                        'success': True,
                        'metadata': result.get('metadata', {}) if isinstance(result, dict) else {}
                    })
                    
                    json_files_created.append(temp_output_file)
                    
                except UnicodeEncodeError as ue:
                    # Handle Unicode encoding errors specifically
                    error_msg = f"Unicode encoding error: {str(ue)}"
                    print(f"Unicode error processing {file_info['file']}: {error_msg}")
                    results.append({
                        'original_file': file_info['file'],
                        'directory': file_info['directory'],
                        'success': False,
                        'error': error_msg
                    })
                except Exception as pe:
                    # Handle other processing errors
                    error_msg = f"Processing error: {str(pe)}"
                    print(f"Error processing {file_info['file']}: {error_msg}")
                    results.append({
                        'original_file': file_info['file'],
                        'directory': file_info['directory'],
                        'success': False,
                        'error': error_msg
                    })
                
            except Exception as e:
                results.append({
                    'original_file': file_info['file'],
                    'directory': file_info['directory'],
                    'success': False,
                    'error': str(e)
                })
        
        # Count schemas in created JSON files
        schema_counts = []
        for json_file in json_files_created:
            if os.path.exists(json_file):
                try:
                    count_result = count_schemas_in_json(json_file)
                    # Extract the total_schemas from the result dictionary
                    schema_count = count_result.get('total_schemas', 0) if isinstance(count_result, dict) else 0
                    schema_counts.append({
                        'json_file': os.path.basename(json_file),
                        'full_path': json_file,
                        'schema_count': schema_count
                    })
                except Exception as e:
                    print(f"Error counting schemas in {json_file}: {str(e)}")
                    schema_counts.append({
                        'json_file': os.path.basename(json_file),
                        'full_path': json_file,
                        'schema_count': 0
                    })
        
        # Perform full schema analysis to create Excel file
        excel_file_path = None
        try:
            # Find schema files for analysis
            schema_files = []
            for file_info in all_files:
                if 'schema' in file_info['file'].lower():
                    schema_files.append(file_info)
            
            if len(schema_files) >= 2:  # Need at least 2 schema files
                # Process schema files to JSON
                bank1_schema_json = None
                bank2_schema_json = None
                
                for schema_file in schema_files:
                    if schema_file['directory'] == 'bank1':
                        base_name = os.path.splitext(schema_file['file'])[0]
                        temp_json = os.path.join(JSON_TEMP_DIR, f"{base_name}_converted.json")
                        try:
                            process_file(schema_file['path'], output_file=temp_json, clean_data=True, include_metadata=True)
                            bank1_schema_json = temp_json
                        except Exception as e:
                            print(f"Error processing bank1 schema: {e}")
                    elif schema_file['directory'] == 'bank2':
                        base_name = os.path.splitext(schema_file['file'])[0]
                        temp_json = os.path.join(JSON_TEMP_DIR, f"{base_name}_converted.json")
                        try:
                            process_file(schema_file['path'], output_file=temp_json, clean_data=True, include_metadata=True)
                            bank2_schema_json = temp_json
                        except Exception as e:
                            print(f"Error processing bank2 schema: {e}")
                
                # Perform schema analysis if we have both schema files
                if bank1_schema_json and bank2_schema_json:
                    print("Starting full schema analysis...")
                    
                    # Send to ChatGPT for schema comparison
                    prompt = """compare the schemas from the two banks provided in the JSON files. For each schema, evaluate all possible combinations and identify the best corresponding schemas based on their descriptions.

MAKE THE MAX NUMBER OF MATCHES AS POSSIBLE, HENCE LEAVE A FEW UN MATCHED SCHEMAS AS POSSIBLE. but do not force connections — if a schema really does not correspond to any other, leave it unmatched.

ALSO DO NOT MATCH A SCHEMA TO ITSELF. OR MATCH A BANK 1 SCHEMA TO A BANK 1 SCHEMA.

No schema should be omitted: every schema from both JSON files must appear in your final output, whether matched or not. HOWEVER, do not repeat any single schema twice

For each schema, provide the following information:
- Schema name
- Description
- Category (e.g., customer, account, transaction)
- Whether it matches with another schema (and which one)

Return the result in a structured JSON format with matched and unmatched schemas."""
                    
                    try:
                        response = send_json_to_chatgpt(bank1_schema_json, prompt, bank2_schema_json)
                    except Exception as api_error:
                        print(f"OpenAI API error: {api_error}")
                        print("Creating fallback Excel file without AI analysis...")
                        response = None
                    
                    if response:
                        # Parse response and create files
                        with open(bank1_schema_json, 'r', encoding='utf-8') as f:
                            bank1_data = json.load(f)
                        with open(bank2_schema_json, 'r', encoding='utf-8') as f:
                            bank2_data = json.load(f)
                        
                        parsed_data = process_chatgpt_schema_analysis(response, bank1_schema_json, bank2_schema_json, JSON_TEMP_DIR)
                        
                        if parsed_data:
                            # Create combined Excel file
                            excel_filename = f"combined_customer_data_{int(time.time())}.xlsx"
                            excel_file_path = os.path.join(os.getcwd(), excel_filename)
                            
                            combined_file = create_combined_customer_data(
                                parsed_data["matched_schemas"], 
                                excel_file_path
                            )
                            
                            if combined_file:
                                excel_file_path = combined_file
                                print(f"Created Excel file: {excel_file_path}")
                            else:
                                excel_file_path = None
                                print("Failed to create Excel file")
                        else:
                            print("Failed to parse ChatGPT response")
                    else:
                        print("No response from ChatGPT - creating fallback Excel file...")
                        # Create a fallback Excel file with all data combined
                        excel_filename = f"combined_customer_data_fallback_{int(time.time())}.xlsx"
                        excel_file_path = os.path.join(os.getcwd(), excel_filename)
                        
                        # Create a simple combined Excel file with all data
                        try:
                            create_fallback_excel_file(json_files_created, excel_file_path)
                            print(f"Created fallback Excel file: {excel_file_path}")
                        except Exception as fallback_error:
                            print(f"Failed to create fallback Excel file: {fallback_error}")
                            excel_file_path = None
                else:
                    print("Missing schema files for analysis")
            else:
                print("Not enough schema files for analysis")
                
        except Exception as e:
            print(f"Error in schema analysis: {e}")
            excel_file_path = None

        return jsonify({
            'success': True,
            'message': f'Processed {len(all_files)} files with main.py',
            'files_processed': len([r for r in results if r['success']]),
            'files_failed': len([r for r in results if not r['success']]),
            'results': results,
            'json_files_created': json_files_created,
            'schema_counts': schema_counts,
            'total_schemas': sum(sc['schema_count'] for sc in schema_counts if isinstance(sc.get('schema_count'), (int, float))),
            'output_directory': os.getcwd(),  # Where the JSON files are created
            'excel_file_path': excel_file_path,  # Path to the created Excel file
            'excel_file_name': os.path.basename(excel_file_path) if excel_file_path else None
        })
        
    except Exception as e:
        return jsonify({'error': f'Error triggering main.py processing: {str(e)}'}), 500
    finally:
        # Restore stdout and stderr
        sys.stdout = old_stdout
        sys.stderr = old_stderr

@app.route('/api/download-excel/<filename>', methods=['GET'])
def download_excel_file(filename):
    """Download the merged Excel file"""
    try:
        # Look for the Excel file in the current directory
        excel_path = os.path.join(os.getcwd(), filename)
        
        if not os.path.exists(excel_path):
            return jsonify({'error': 'Excel file not found'}), 404
        
        # Send the file
        return send_file(excel_path, as_attachment=True, download_name=filename)
        
    except Exception as e:
        return jsonify({'error': f'Error downloading Excel file: {str(e)}'}), 500

@app.route('/api/cleanup-json-files', methods=['POST'])
def cleanup_json_files():
    """Clean up all files (JSON and uploaded files)"""
    try:
        success, cleaned_files, failed_files = cleanup_all_files()
        
        response_data = {
            'success': success,
            'cleaned_files': cleaned_files,
            'failed_files': failed_files,
            'total_cleaned': len(cleaned_files),
            'total_failed': len(failed_files)
        }
        
        if success:
            response_data['message'] = f'All files cleaned up successfully! Cleaned {len(cleaned_files)} files.'
        else:
            response_data['message'] = f'Partial cleanup completed. Cleaned {len(cleaned_files)} files, {len(failed_files)} files failed.'
        
        return jsonify(response_data)
    except Exception as e:
        return jsonify({'error': f'Error cleaning up files: {str(e)}'}), 500

def compare_schemas_with_main():
    """Compare schemas using main.py functionality"""
    try:
        # Import main.py functions
        sys.path.append(os.path.dirname(os.path.abspath(__file__)))
        from main import send_json_to_chatgpt, parse_chatgpt_response, create_schema_json_files
        
        # Get all schema files from json_storage
        schema_files = []
        for filename in os.listdir(UPLOAD_STORAGE_DIR):
            if filename.endswith('.json') and 'schema' in filename.lower():
                schema_files.append(filename)
        
        if len(schema_files) < 2:
            return jsonify({'error': 'Need at least 2 schema files for comparison'}), 400
        
        # Use the first two schema files
        file1_path = os.path.join(UPLOAD_STORAGE_DIR, schema_files[0])
        file2_path = os.path.join(UPLOAD_STORAGE_DIR, schema_files[1])
        
        # Send to ChatGPT for comparison (this would require OpenAI API key)
        prompt = """
        Compare the schemas from the two banks provided in the JSON files. 
        For each schema, evaluate all possible combinations and identify the best corresponding schemas based on their descriptions.
        
        Make the maximum number of matches as possible, but do not force connections.
        Do not match a schema to itself or match schemas from the same bank.
        
        Provide your analysis in a structured format.
        """
        
        # Note: This would require OpenAI API key setup
        # response = send_json_to_chatgpt(file1_path, prompt)
        
        return jsonify({
            'success': True,
            'message': 'Schema comparison functionality ready',
            'file1': schema_files[0],
            'file2': schema_files[1],
            'note': 'OpenAI API integration requires API key configuration'
        })
    
    except Exception as e:
        return jsonify({'error': f'Error comparing schemas: {str(e)}'}), 500

def process_file_with_main(file_path):
    """Process a file using main.py functionality"""
    try:
        # Import main.py functions
        sys.path.append(os.path.dirname(os.path.abspath(__file__)))
        from main import process_file
        
        # Process the file
        result = process_file(file_path)
        
        return jsonify({
            'success': True,
            'message': 'File processed successfully',
            'result': result
        })
    
    except Exception as e:
        return jsonify({'error': f'Error processing file: {str(e)}'}), 500

if __name__ == '__main__':
    print("Starting Bridgette Backend Server...")
    print(f"Backend will be available at: http://{app.config['HOST']}:5001")
    print("API Endpoints:")
    print("  POST /api/process-files - Save uploaded files for main.py access")
    print("    - ?schema=true for schema files")
    print("    - No parameter for normal files")
    print("  GET  /api/health - Health check")
    print("  GET  /api/uploaded-files - List all saved uploaded files")
    print("  GET  /api/uploaded-files/<filename> - Get specific uploaded file info")
    print("  DELETE /api/uploaded-files/<filename> - Delete uploaded file")
    print("")
    print("  POST /api/process-with-main - Process files using main.py functionality")
    print("  POST /api/trigger-main-processing - Process all uploaded files with main.py")
    print("  POST /api/cleanup-json-files - Clean up temporary JSON files")
    print("Supported formats: CSV, Excel (.xlsx, .xls)")
    print("Features:")
    print("  - Saves original files in uploaded_files/ directory")
    print("  - Preserves filenames and metadata")
    print("  - Files accessible to main.py for processing")
    print("  - Optimized for large files (up to 50MB)")
    print(f"Environment: {'Development' if app.config['DEBUG'] else 'Production'}")
    app.run(debug=app.config['DEBUG'], host=app.config['HOST'], port=5001)